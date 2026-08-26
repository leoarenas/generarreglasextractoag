import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable


CANONICAL_FIELDS = [
    "fecha_movimiento",
    "fecha_valor",
    "descripcion",
    "importe",
    "tipo_movimiento",
    "debito",
    "credito",
    "saldo",
    "referencia",
    "moneda",
    "cuenta",
    "codigo_movimiento",
    "contraparte",
    "id_movimiento",
]

REQUIRED_IDENTITY_FIELDS = ("bank_name", "account_number", "account_type", "currency")

ALIASES = {
    "fecha_movimiento": ("fecha", "fecha movimiento", "f movimiento", "fecha operacion"),
    "fecha_valor": ("fecha valor", "f valor"),
    "descripcion": ("concepto", "descripcion", "detalle", "movimiento"),
    "importe": ("importe", "monto", "valor"),
    "tipo_movimiento": ("tipo movimiento", "tipo", "debito credito"),
    "debito": ("debito", "debitos", "egreso", "egresos", "debe"),
    "credito": ("credito", "creditos", "ingreso", "ingresos", "haber"),
    "saldo": ("saldo", "balance"),
    "referencia": ("referencia", "nro referencia", "numero referencia", "comprobante"),
    "moneda": ("moneda", "currency"),
    "cuenta": ("cuenta", "numero cuenta", "nro cuenta"),
    "codigo_movimiento": ("causal", "codigo movimiento", "cod movimiento"),
    "contraparte": ("contraparte", "ordenante", "beneficiario"),
    "id_movimiento": ("id movimiento", "id operacion", "identificador"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def normalize_label(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalize_label(value)).strip("_") or "perfil"


def infer_mapping(headers: list[str]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    used: set[str] = set()
    for header in headers:
        normalized = normalize_label(header)
        match = None
        for field, aliases in ALIASES.items():
            if field not in used and normalized in aliases:
                match = field
                used.add(field)
                break
        result[header] = match
    return result


def validate_mapping(mapping: dict[str, str | None]) -> None:
    fields = {field for field in mapping.values() if field}
    missing = [field for field in ("fecha_movimiento", "descripcion") if field not in fields]
    has_amount = "importe" in fields or {"debito", "credito"}.issubset(fields)
    if missing or not has_amount:
        details = []
        if missing:
            details.append("faltan " + ", ".join(missing))
        if not has_amount:
            details.append("se requiere importe o el par debito/credito")
        raise ValueError("Diccionario de columnas incompleto: " + "; ".join(details) + ".")
    mapped = [field for field in mapping.values() if field]
    duplicates = sorted({field for field in mapped if mapped.count(field) > 1})
    if duplicates:
        raise ValueError("Campos normalizados repetidos: " + ", ".join(duplicates))


def read_tabular_file(
    path: Path,
    sheet_name: str | None = None,
    pdf_password: str | None = None,
) -> tuple[list[list[str]], str | None]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = path.read_text(encoding="utf-8-sig")
        dialect = csv.Sniffer().sniff(raw[:8192], delimiters=",;\t|")
        return [[_text(cell) for cell in row] for row in csv.reader(raw.splitlines(), dialect)], None
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Instale openpyxl para leer extractos XLSX.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        selected = sheet_name or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"La hoja '{selected}' no existe. Disponibles: {workbook.sheetnames}")
        worksheet = workbook[selected]
        return [[_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)], selected
    if suffix == ".pdf":
        try:
            import pdfplumber
        except ImportError as exc:
            raise RuntimeError("Instale pdfplumber para leer extractos PDF.") from exc
        extracted_rows: list[list[str]] = []
        with pdfplumber.open(path, password=pdf_password) as document:
            for page in document.pages:
                tables = page.extract_tables() or []
                if not tables:
                    continue
                table = max(tables, key=lambda candidate: len(candidate or []))
                extracted_rows.extend(
                    [[_text(cell) for cell in row] for row in table if row]
                )
        if not extracted_rows:
            raise ValueError(
                "No se detectaron tablas en el PDF. Si el extracto esta escaneado, "
                "conviertalo a un PDF con texto/OCR o solicite al banco CSV/XLSX."
            )
        return extracted_rows, None
    raise ValueError("Formato no soportado. Use CSV, XLSX, XLSM o PDF.")


def read_positioned_pdf_account(
    path: Path, account_hint: str, pdf_password: str | None = None
) -> tuple[list[str], list[list[str]]]:
    """Extrae movimientos de PDFs bancarios con columnas visuales pero sin tabla embebida."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("Instale pdfplumber para leer extractos PDF.") from exc

    headers = ["FECHA", "DESCRIPCION", "REFERENCIA", "DEBITOS", "CREDITOS", "SALDO"]
    movements: list[list[str]] = []
    active_account = False
    with pdfplumber.open(path, password=pdf_password) as document:
        for page in document.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            grouped: list[list[dict[str, Any]]] = []
            for word in sorted(words, key=lambda item: (round(item["top"], 1), item["x0"])):
                if not grouped or abs(grouped[-1][0]["top"] - word["top"]) > 1:
                    grouped.append([word])
                else:
                    grouped[-1].append(word)

            for line_words in grouped:
                line_words.sort(key=lambda item: item["x0"])
                line = " ".join(item["text"] for item in line_words)
                normalized_line = normalize_label(line)
                if normalized_line.startswith("cuenta ") and " nro " in f" {normalized_line} ":
                    compact_hint = re.sub(r"[^a-z0-9]", "", normalize_label(account_hint))
                    compact_line = re.sub(r"[^a-z0-9]", "", normalized_line)
                    active_account = compact_hint in compact_line
                    continue
                if not active_account or not re.match(r"^\d{2}/\d{2}/\d{2}$", line_words[0]["text"]):
                    continue

                columns: list[list[str]] = [[] for _ in headers]
                for word in line_words:
                    x = float(word["x0"])
                    if x < 65:
                        index = 0
                    elif x < 255:
                        index = 1
                    elif x < 330:
                        index = 2
                    elif x < 410:
                        index = 3
                    elif x < 495:
                        index = 4
                    else:
                        index = 5
                    columns[index].append(word["text"])
                movements.append([" ".join(values).strip() for values in columns])

    if not movements:
        raise ValueError(
            f"No se encontraron movimientos para la cuenta identificada por '{account_hint}'."
        )
    return headers, movements


def detect_header_row(rows: list[list[str]], scan_limit: int = 25) -> int:
    candidates = []
    known_aliases = {alias for aliases in ALIASES.values() for alias in aliases}
    for index, row in enumerate(rows[:scan_limit]):
        values = [normalize_label(cell) for cell in row if _text(cell)]
        alias_hits = sum(value in known_aliases for value in values)
        candidates.append((alias_hits, len(set(values)), index))
    if not candidates or max(candidates)[:2] == (0, 0):
        raise ValueError("No se pudo detectar una fila de encabezados.")
    return max(candidates)[2]


def extract_table(rows: list[list[str]], header_row: int) -> tuple[list[str], list[list[str]]]:
    raw_header = rows[header_row]
    indexes = [index for index, value in enumerate(raw_header) if _text(value)]
    headers = [_text(raw_header[index]) for index in indexes]
    if len(headers) != len(set(headers)):
        raise ValueError("El extracto contiene encabezados duplicados.")
    data = []
    for row in rows[header_row + 1 :]:
        selected = [_text(row[index]) if index < len(row) else "" for index in indexes]
        if any(selected):
            data.append(selected)
    if not data:
        raise ValueError("No se encontraron movimientos debajo de los encabezados.")
    return headers, data


def parse_amount(value: str, decimal_separator: str, thousands_separator: str) -> Decimal | None:
    text = _text(value).replace(" ", "")
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(thousands_separator, "")
    if decimal_separator != ".":
        text = text.replace(decimal_separator, ".")
    text = re.sub(r"[^0-9.\-+]", "", text)
    try:
        amount = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Importe invalido: {value}") from exc
    return -amount if negative else amount


def normalize_rows(headers: list[str], data: list[list[str]], profile: dict[str, Any]) -> list[dict[str, Any]]:
    mapping = profile["columns"]
    rules = profile.get("normalization", {})
    decimal_separator = rules.get("decimal_separator", ",")
    thousands_separator = rules.get("thousands_separator", ".")
    account = profile["account"]
    normalized = []
    amount_fields = {"importe", "debito", "credito", "saldo"}
    for row in data:
        source = dict(zip(headers, row))
        item: dict[str, Any] = {field: None for field in CANONICAL_FIELDS}
        for source_name, field in mapping.items():
            if field and source_name in source:
                value: Any = source[source_name]
                if field in amount_fields:
                    parsed = parse_amount(value, decimal_separator, thousands_separator)
                    value = float(parsed) if parsed is not None else None
                item[field] = value or None
        item["moneda"] = item["moneda"] or account["currency"]
        item["cuenta"] = item["cuenta"] or account["number"]
        if item["importe"] is None:
            debit = item["debito"] or 0
            credit = item["credito"] or 0
            item["importe"] = credit - debit
        if not item["tipo_movimiento"]:
            item["tipo_movimiento"] = "credito" if item["importe"] >= 0 else "debito"
        normalized.append(item)
    return normalized


def save_profile(profile: dict[str, Any], profiles_dir: Path) -> Path:
    for field in REQUIRED_IDENTITY_FIELDS:
        if not _text(profile["account"].get(field.replace("account_", ""))):
            raise ValueError(f"Dato obligatorio faltante: {field}")
    validate_mapping(profile["columns"])
    profiles_dir.mkdir(parents=True, exist_ok=True)
    base = slugify(profile["profile_name"])
    version = 1
    while (profiles_dir / f"{base}_v{version}.json").exists():
        version += 1
    profile["version"] = version
    target = profiles_dir / f"{base}_v{version}.json"
    target.write_text(json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8")
    return target


def load_profile(path: Path) -> dict[str, Any]:
    profile = json.loads(path.read_text(encoding="utf-8"))
    validate_mapping(profile["columns"])
    return profile


def _required_input(label: str, input_fn: Callable[[str], str]) -> str:
    while True:
        value = input_fn(f"{label}: ").strip()
        if value:
            return value
        print("Este dato es obligatorio.")


def run_setup_wizard(
    sample_file: Path | None = None,
    profiles_dir: Path = Path("config/profiles"),
    input_fn: Callable[[str], str] = input,
) -> Path:
    print("\nConfiguracion inicial del agente de conciliacion\n")
    bank_name = _required_input("Nombre del banco", input_fn)
    account_number = _required_input(
        "Numero, alias o ultimos 4 caracteres de la cuenta", input_fn
    )
    account_type = _required_input(
        "Tipo de cuenta (ejemplos: Cuenta corriente, Caja de ahorro, Cuenta recaudadora)",
        input_fn,
    )
    currency = _required_input("Moneda (por ejemplo ARS o USD)", input_fn).upper()
    sample_file = sample_file or Path(
        _required_input(
            "Ruta del extracto de muestra o del archivo adjunto (CSV, XLSX, XLSM o PDF)",
            input_fn,
        )
    )
    if not sample_file.exists():
        raise FileNotFoundError(f"No existe el archivo: {sample_file}")

    pdf_password = None
    if sample_file.suffix.lower() == ".pdf":
        pdf_password = input_fn(
            "Contraseña del PDF (dejar vacio si no tiene): "
        ).strip() or None
        try:
            headers, data = read_positioned_pdf_account(
                sample_file, account_number, pdf_password
            )
            rows = [headers] + data
            selected_sheet = None
        except ValueError:
            rows, selected_sheet = read_tabular_file(
                sample_file, pdf_password=pdf_password
            )
    else:
        rows, selected_sheet = read_tabular_file(sample_file)
    header_index = detect_header_row(rows)
    headers, data = extract_table(rows, header_index)
    proposed = infer_mapping(headers)

    print(f"\nEncabezados detectados en la fila {header_index + 1}:")
    print("Campos disponibles: " + ", ".join(CANONICAL_FIELDS))
    mapping: dict[str, str | None] = {}
    for header in headers:
        suggestion = proposed[header] or "ignorar"
        answer = input_fn(f"'{header}' -> [{suggestion}]: ").strip()
        value = answer or suggestion
        mapping[header] = None if value.lower() in {"ignorar", "ignore", "-"} else value
    validate_mapping(mapping)

    decimal_separator = input_fn("Separador decimal [,]: ").strip() or ","
    thousands_separator = input_fn("Separador de miles [.]: ").strip() or "."
    profile_name = input_fn(
        f"Nombre del perfil [{bank_name}_{account_number}_{currency}]: "
    ).strip() or f"{bank_name}_{account_number}_{currency}"
    profile = {
        "profile_name": profile_name,
        "created_at": datetime.now().astimezone().isoformat(),
        "account": {
            "bank_name": bank_name,
            "bank_code": slugify(bank_name),
            "number": account_number,
            "type": account_type,
            "currency": currency,
        },
        "format": {
            "file_type": sample_file.suffix.lower().lstrip("."),
            "sheet_name": selected_sheet,
            "header_row": header_index + 1,
            "parser": "positioned_pdf_account" if sample_file.suffix.lower() == ".pdf" else "table",
        },
        "columns": mapping,
        "normalization": {
            "decimal_separator": decimal_separator,
            "thousands_separator": thousands_separator,
        },
    }
    preview = normalize_rows(headers, data[:10], profile)
    print("\nVista previa normalizada:")
    print(json.dumps(preview, indent=2, ensure_ascii=False))
    confirmation = _required_input("Guardar esta configuracion? (si/no)", input_fn).lower()
    if confirmation not in {"si", "sí", "s", "yes", "y"}:
        raise RuntimeError("Configuracion cancelada por el usuario.")
    return save_profile(profile, profiles_dir)


def normalized_file_rows(
    source_file: Path,
    profile: dict[str, Any],
    pdf_password: str | None = None,
) -> list[list[str]]:
    if profile.get("format", {}).get("parser") == "positioned_pdf_account":
        headers, data = read_positioned_pdf_account(
            source_file, profile["account"]["number"], pdf_password
        )
    else:
        rows, _ = read_tabular_file(
            source_file,
            profile.get("format", {}).get("sheet_name"),
            pdf_password,
        )
        expected_header_row = int(profile["format"]["header_row"]) - 1
        headers, data = extract_table(rows, expected_header_row)
    expected = set(profile["columns"])
    found = set(headers)
    if missing := sorted(expected - found):
        raise ValueError("El formato cambio. Columnas esperadas ausentes: " + ", ".join(missing))
    if new_columns := sorted(found - expected):
        raise ValueError("El formato cambio. Columnas nuevas detectadas: " + ", ".join(new_columns))
    normalized = normalize_rows(headers, data, profile)
    return [CANONICAL_FIELDS] + [[_text(row[field]) for field in CANONICAL_FIELDS] for row in normalized]
